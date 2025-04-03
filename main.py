import os
import re
import base64
import json
import requests
import feedparser
import urllib.parse
import concurrent.futures
from datetime import datetime, timedelta
from dotenv import load_dotenv
from tabulate import tabulate
import openpyxl
from openpyxl.styles import Alignment, Font
from bs4 import BeautifulSoup
import time
import traceback
import logging
from tqdm import tqdm
import tiktoken  # 토큰 계산을 위한 OpenAI 라이브러리

# Gmail API 관련 모듈
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from googleapiclient.discovery import build
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request


# OpenAI API 사용량 추적 및 비용 계산 클래스
class OpenAIUsageTracker:
    def __init__(self, exchange_rate=1480):
        """
        OpenAI API 사용량 추적 및 비용 계산 클래스 초기화

        Args:
            exchange_rate: 달러 대 원화 환율 (기본값: 1480원)
        """
        self.exchange_rate = exchange_rate
        self.total_input_tokens = 0
        self.total_output_tokens = 0
        self.total_requests = 0

        # GPT-4o-mini 모델 가격 설정 (달러/100만 토큰)
        self.input_price_per_million = 0.15  # 100만 토큰당 0.15 달러
        self.output_price_per_million = 0.60  # 100만 토큰당 0.60 달러

        # 토큰 계산을 위한 인코더 초기화
        try:
            self.encoder = tiktoken.encoding_for_model("gpt-4o-mini")
        except:
            self.encoder = tiktoken.get_encoding("cl100k_base")  # 대체 인코더
            logging.warning(
                "gpt-4o-mini 인코더를 로드할 수 없어 cl100k_base를 대신 사용합니다."
            )

    def count_tokens(self, text):
        """문자열의 토큰 수를 계산합니다."""
        if not text:
            return 0
        return len(self.encoder.encode(text))

    def track_request(self, messages):
        """API 요청의 토큰 사용량을 추적합니다."""
        input_tokens = sum(
            self.count_tokens(msg.get("content", "")) for msg in messages
        )
        self.total_input_tokens += input_tokens
        self.total_requests += 1
        return input_tokens

    def track_response(self, response_text):
        """API 응답의 토큰 사용량을 추적합니다."""
        output_tokens = self.count_tokens(response_text)
        self.total_output_tokens += output_tokens
        return output_tokens

    def calculate_cost(self):
        """현재까지의 API 사용량 비용을 계산합니다."""
        # 입력 토큰 비용 (달러)
        input_cost_usd = (
            self.total_input_tokens / 1000000
        ) * self.input_price_per_million
        # 출력 토큰 비용 (달러)
        output_cost_usd = (
            self.total_output_tokens / 1000000
        ) * self.output_price_per_million
        # 총 비용 (달러)
        total_cost_usd = input_cost_usd + output_cost_usd

        # 원화로 변환
        input_cost_krw = input_cost_usd * self.exchange_rate
        output_cost_krw = output_cost_usd * self.exchange_rate
        total_cost_krw = total_cost_usd * self.exchange_rate

        return {
            "input_tokens": self.total_input_tokens,
            "output_tokens": self.total_output_tokens,
            "total_tokens": self.total_input_tokens + self.total_output_tokens,
            "total_requests": self.total_requests,
            "input_cost_usd": input_cost_usd,
            "output_cost_usd": output_cost_usd,
            "total_cost_usd": total_cost_usd,
            "input_cost_krw": input_cost_krw,
            "output_cost_krw": output_cost_krw,
            "total_cost_krw": total_cost_krw,
        }

    def log_usage(self):
        """현재까지의 API 사용량 및 비용을 로그에 기록합니다."""
        cost_data = self.calculate_cost()
        logging.info("===== OpenAI API 사용량 및 비용 =====")
        logging.info(f"총 요청 수: {cost_data['total_requests']:,}회")
        logging.info(
            f"입력 토큰: {cost_data['input_tokens']:,}개 (${cost_data['input_cost_usd']:.6f}, ₩{cost_data['input_cost_krw']:.2f})"
        )
        logging.info(
            f"출력 토큰: {cost_data['output_tokens']:,}개 (${cost_data['output_cost_usd']:.6f}, ₩{cost_data['output_cost_krw']:.2f})"
        )
        logging.info(
            f"전체 토큰: {cost_data['total_tokens']:,}개 (${cost_data['total_cost_usd']:.6f}, ₩{cost_data['total_cost_krw']:.2f})"
        )
        logging.info("====================================")
        return cost_data


# ANSI 색상 코드 기반의 커스텀 Formatter 정의
class ColoredFormatter(logging.Formatter):
    RESET = "\x1b[0m"
    COLORS = {
        logging.DEBUG: "\x1b[37;20m",  # 회색
        logging.INFO: "\x1b[32;20m",  # 초록색
        logging.WARNING: "\x1b[33;20m",  # 노란색
        logging.ERROR: "\x1b[31;20m",  # 빨간색
        logging.CRITICAL: "\x1b[31;1m",  # 진한 빨간색
    }

    def format(self, record):
        log_color = self.COLORS.get(record.levelno, self.RESET)
        formatter = logging.Formatter(
            f"{log_color}%(asctime)s - %(levelname)s - %(message)s{self.RESET}",
            "%Y-%m-%d %H:%M:%S",
        )
        return formatter.format(record)


# 로그 파일명을 타임스탬프를 포함해서 생성
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
log_filename = f"ynanews-{timestamp}.log"

# 로그 폴더 생성 (없는 경우)
log_dir = "logs"
if not os.path.exists(log_dir):
    os.makedirs(log_dir)

# 로그 경로 변수 설정 (모듈 레벨 변수)
log_path = os.path.join(log_dir, log_filename)

# FileHandler를 이용해서 로그를 파일에 기록하도록 설정 (인코딩 'utf-8' 지정)
file_handler = logging.FileHandler(log_path, encoding="utf-8")
file_handler.setFormatter(ColoredFormatter())
file_handler.setLevel(logging.ERROR)  # 로그 파일에는 ERROR 수준만 기록

# 루트 로거 설정
logger = logging.getLogger()
logger.handlers = [file_handler]
logger.setLevel(logging.INFO)  # 로깅 레벨은 INFO로 유지

# 콘솔 출력 핸들러 추가 (모든 INFO 이상 레벨 표시)
console_handler = logging.StreamHandler()
console_handler.setFormatter(logging.Formatter("%(levelname)s: %(message)s"))
console_handler.setLevel(logging.INFO)  # 콘솔에는 INFO 수준부터 모두 출력
logger.addHandler(console_handler)

logging.info(f"로그 파일 경로: {log_path} (ERROR 레벨 로그만 기록됨)")
logging.info("콘솔 출력: INFO 레벨 이상의 로그가 출력됩니다.")


def main():
    logging.info("프로그램 시작")
    logging.info(f"작업 폴더: {os.getcwd()}")
    logging.info(f"로그 폴더: {os.path.abspath('logs')}")
    logging.info(f"출력 폴더: {os.path.abspath('output')}")

    # 0. .env 파일에서 환경 변수 로드
    load_dotenv()
    OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
    keywords_env = os.getenv("SEARCH_KEYWORDS", "")
    keyword_list = (
        [kw.strip() for kw in keywords_env.split(",")]
        if keywords_env
        else ["ICT", "AI"]
    )

    # OpenAI API 사용량 추적기 초기화 (환율: 1$=1480원)
    usage_tracker = OpenAIUsageTracker(exchange_rate=1480)

    logging.info("OpenAI API Key Loaded: %s", bool(OPENAI_API_KEY))
    logging.info("검색 키워드: %s", keyword_list)

    # 1. YNA RSS 피드에서 기사 수집
    yna_rss_url = "https://www.yna.co.kr/rss/news.xml"
    rss_data = feedparser.parse(yna_rss_url)
    articles = rss_data["entries"]
    logging.info("전체 기사 개수: %d", len(articles))
    if articles:
        logging.info("첫 번째 기사 제목: %s", articles[0].title)

    # 2. 최근 2일 이내 기사 필터링
    two_days_ago = datetime.now() - timedelta(days=2)
    recent_articles = []
    for art in articles:
        try:
            pub_date = datetime(*art.published_parsed[:6])
        except Exception:
            continue
        if pub_date >= two_days_ago:
            art.pub_date_obj = pub_date
            recent_articles.append(art)
    logging.info("최근 2일 이내 기사 개수: %d", len(recent_articles))

    # 2.5. 실제 URL에서 기사 본문을 추출하는 함수
    def get_full_article_content(url):
        logging.info("Fetching article content from: %s", url)
        try:
            headers = {"User-Agent": "Mozilla/5.0"}
            response = requests.get(url, headers=headers, timeout=10)
            response.raise_for_status()
            soup = BeautifulSoup(response.text, "html.parser")

            # 1) 'div.story-news' 컨테이너 내의 <p> 태그들만 추출
            container = soup.select_one("div.story-news")
            if container:
                paragraphs = container.find_all("p")
                if paragraphs:
                    content = "\n".join(p.get_text(strip=True) for p in paragraphs)
                    if len(content) > 200:
                        logging.debug(
                            "Extracted article content from paragraphs in 'div.story-news'"
                        )
                        return content
                text = container.get_text(separator="\n").strip()
                if len(text) > 200:
                    return text

            # 2) 'div.article-body' 또는 'article' 선택자 사용
            for sel in ["div.article-body", "article"]:
                container = soup.select_one(sel)
                if container:
                    paragraphs = container.find_all("p")
                    if paragraphs:
                        content = "\n".join(p.get_text(strip=True) for p in paragraphs)
                        if len(content) > 200:
                            logging.debug(
                                "Extracted article content from paragraphs in '%s'", sel
                            )
                            return content
                    text = container.get_text(separator="\n").strip()
                    if len(text) > 200:
                        return text

            # 3) fallback: 전체 페이지 텍스트 반환
            logging.debug("No specific container found, returning full page text")
            return soup.get_text(separator="\n").strip()
        except Exception as e:
            logging.error(
                "Error fetching article content from URL: %s - %s", url, str(e)
            )
            traceback.print_exc()
            return ""

    # 3. OpenAI API를 이용해 기사 본문 요약 생성 함수
    def summarize_text(text):
        truncated_text = text if len(text) < 1000 else text[:1000]
        prompt = f"다음 뉴스 기사 본문을 간략하게 요약해줘. 요약은 3문장 이내로 작성해줘:\n\n{truncated_text}"
        messages = [
            {
                "role": "system",
                "content": "You are an assistant that summarizes news articles in Korean.",
            },
            {"role": "user", "content": prompt},
        ]

        # 요청 토큰 추적
        input_tokens = usage_tracker.track_request(messages)
        logging.debug(f"요약 요청 입력 토큰: {input_tokens}개")

        data = {
            "model": "gpt-4o-mini",
            "messages": messages,
            "temperature": 0.3,
        }
        req_headers = {
            "Authorization": f"Bearer {OPENAI_API_KEY}",
            "Content-Type": "application/json",
        }
        response = requests.post(
            "https://api.openai.com/v1/chat/completions", json=data, headers=req_headers
        )
        if response.status_code != 200:
            logging.error(
                "OpenAI API error in summarization: %d - %s",
                response.status_code,
                response.text,
            )
            return ""
        result = response.json()
        summary = result["choices"][0]["message"]["content"].strip()

        # 응답 토큰 추적
        output_tokens = usage_tracker.track_response(summary)
        logging.debug(f"요약 응답 출력 토큰: {output_tokens}개")

        return summary

    # 4. 키워드 검사 함수
    def is_relevant(art_text):
        return any(keyword in art_text for keyword in keyword_list)

    # 5. 초기 필터링: RSS의 title과 summary로 기사 선별
    pre_filtered_articles = [
        art
        for art in recent_articles
        if is_relevant(art.title) or is_relevant(art.get("summary", ""))
    ]
    logging.info("초기 필터링 후 기사 개수 (RSS 기준): %d", len(pre_filtered_articles))

    # 6. 동시 처리로 실제 본문 가져오기 및 재필터링
    def fetch_and_filter(art, pbar):
        full_content = get_full_article_content(art.link)
        pbar.update(1)
        if is_relevant(art.title) or is_relevant(full_content):
            return {
                "title": art.title,
                "content": full_content,
                "url": art.link,
                "pubdate": art.pub_date_obj,
            }
        return None

    filtered_articles = []
    with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
        futures = {}
        with tqdm(
            total=len(pre_filtered_articles),
            desc="Fetching and filtering articles",
            colour="green",
        ) as pbar:
            # 작업 제출 및 콜백 설정
            for art in pre_filtered_articles:
                futures[executor.submit(fetch_and_filter, art, pbar)] = art

            # 작업 결과 수집
            for future in concurrent.futures.as_completed(futures):
                result = future.result()
                if result:
                    filtered_articles.append(result)

    logging.info("최종 키워드 필터링 후 기사 개수: %d", len(filtered_articles))

    # 7. 각 기사에 대해 본문 요약 생성 (동적 진행률 표시)
    with tqdm(
        total=len(filtered_articles), desc="Summarizing articles", colour="blue"
    ) as pbar:
        for art in filtered_articles:
            logging.info("Summarizing article: %s", art["title"])
            art["article_summary"] = summarize_text(art["content"])
            pbar.update(1)

    # 8. 두 기사 제목 간 유사도 측정 함수 (0 ~ 100)
    def calculate_similarity(title1, title2):
        prompt = (
            f"두 기사 제목의 유사도를 0에서 100까지의 숫자로 측정해줘.\n"
            f'기사 제목1: "{title1}"\n기사 제목2: "{title2}"\n'
            "답변은 순수한 숫자(예: 95)만 출력해줘."
        )
        messages = [
            {
                "role": "system",
                "content": "You are an assistant that calculates similarity percentage between two texts.",
            },
            {"role": "user", "content": prompt},
        ]

        # 요청 토큰 추적
        input_tokens = usage_tracker.track_request(messages)
        logging.debug(f"유사도 계산 요청 입력 토큰: {input_tokens}개")

        data = {
            "model": "gpt-4o-mini",
            "messages": messages,
            "temperature": 0.0,
        }
        req_headers = {
            "Authorization": f"Bearer {OPENAI_API_KEY}",
            "Content-Type": "application/json",
        }

        max_retries = 3
        retry_delay = 2  # 초

        for attempt in range(max_retries):
            try:
                response = requests.post(
                    "https://api.openai.com/v1/chat/completions",
                    json=data,
                    headers=req_headers,
                    timeout=30,
                )
                if response.status_code != 200:
                    logging.error(
                        "OpenAI API 오류 (시도 %d/%d): %d - %s",
                        attempt + 1,
                        max_retries,
                        response.status_code,
                        response.text,
                    )
                    if attempt < max_retries - 1:
                        logging.info("%d초 후 재시도...", retry_delay)
                        time.sleep(retry_delay)
                        continue

                result = response.json()
                answer = result["choices"][0]["message"]["content"].strip()

                # 응답 토큰 추적
                output_tokens = usage_tracker.track_response(answer)
                logging.debug(f"유사도 계산 응답 출력 토큰: {output_tokens}개")

                match = re.search(r"(\d+(\.\d+)?)", answer)
                if match:
                    similarity = float(match.group(1))
                    logging.info(
                        "유사도 계산: '%s' vs '%s' -> %.1f%%",
                        title1,
                        title2,
                        similarity,
                    )
                    return similarity
                return 0.0

            except requests.exceptions.RequestException as e:
                logging.error(
                    "네트워크 오류 (시도 %d/%d): %s", attempt + 1, max_retries, str(e)
                )
                traceback.print_exc()
                if attempt < max_retries - 1:
                    logging.info("%d초 후 재시도...", retry_delay)
                    time.sleep(retry_delay)
                else:
                    logging.error("최대 재시도 횟수 초과, 유사도 계산 실패")
                    return 0.0
        return 0.0

    # 9. 구글 뉴스에서 동일 기사(유사도 70% 이상)를 최대 3개까지 찾는 함수
    def get_google_duplicates(article, pbar=None):
        query = article["title"]
        encoded_query = urllib.parse.quote(query)
        google_news_url = f"https://news.google.com/rss/search?q={encoded_query}&hl=ko&gl=KR&ceid=KR:ko"
        logging.info("[DEBUG] 구글 뉴스 검색 URL: %s", google_news_url)
        search_results = feedparser.parse(google_news_url)
        duplicates = []

        # 내부 진행률 바 생성 (leave=False로 설정하여 완료 후 사라지게 함)
        with tqdm(
            total=min(len(search_results["entries"]), 10),
            desc=f"검색: {article['title'][:20]}...",
            colour="magenta",
            leave=False,
        ) as inner_pbar:
            for entry in search_results["entries"][:10]:  # 최대 10개 항목만 검사
                sim = calculate_similarity(article["title"], entry.title)
                inner_pbar.update(1)
                if sim >= 70:
                    duplicates.append(
                        {"title": entry.title, "url": entry.link, "similarity": sim}
                    )
                    if len(duplicates) >= 3:  # 최대 3개까지만 검색
                        break

        # 외부 진행률 바 업데이트
        if pbar:
            pbar.update(1)

        return duplicates

    # 10. 각 기사에 대해 구글 뉴스 중복 기사 목록 추가 (동적 진행률 표시)
    with tqdm(
        total=len(filtered_articles), desc="Finding duplicates", colour="yellow"
    ) as pbar:
        for art in filtered_articles:
            logging.info("[DEBUG] 처리 중 기사: %s", art["title"])
            art["duplicates"] = get_google_duplicates(art, pbar)

    # 11. 최종 결과 준비 (엑셀 파일 저장을 위한 데이터 생성)
    table_headers = ["기사 제목", "기사 본문", "본문 요약", "기사 URL", "발행일"]
    for i in range(3):
        table_headers.extend(
            [f"중복기사 {i+1} 제목", f"중복기사 {i+1} URL", f"유사율 {i+1} (%)"]
        )

    rows = []
    for art in filtered_articles:
        row = [
            art["title"],
            art["content"],
            art.get("article_summary", ""),
            art["url"],
            art["pubdate"].strftime("%Y-%m-%d %H:%M:%S"),
        ]
        for i in range(3):
            if i < len(art.get("duplicates", [])):
                dup = art["duplicates"][i]
                row.extend([dup["title"], dup["url"], f"{dup['similarity']:.1f}"])
            else:
                row.extend(["", "", ""])
        rows.append(row)

    logging.info("최종 결과 준비 완료")
    # 콘솔에 엑셀 파일 내용을 출력하는 부분은 삭제됨

    # 12. 결과를 엑셀 파일로 저장하는 함수
    def save_to_excel(data, headers, filename=None):
        # output 폴더 생성 (없는 경우)
        output_dir = "output"
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            logging.info(f"'{output_dir}' 폴더 생성 완료")

        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = os.path.join(output_dir, f"output_{timestamp}.xlsx")
        elif not os.path.isabs(filename):
            # 상대 경로인 경우 output 폴더 아래에 저장
            filename = os.path.join(output_dir, os.path.basename(filename))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "뉴스 데이터"
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(wrap_text=True)
        for column in ws.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = min(len(str(cell.value)), 50)
                except Exception:
                    pass
            adjusted_width = max_length + 2
            ws.column_dimensions[column_letter].width = adjusted_width
        wb.save(filename)
        logging.info("엑셀 파일 저장 완료: %s", filename)
        return filename

    # 13. 최종 결과를 엑셀로 저장하고, 파일명을 반환
    output_filename = save_to_excel(rows, table_headers)

    # 프로그램 종료 전 API 사용량 및 비용 로깅
    usage_data = usage_tracker.log_usage()

    # 토큰 파일 존재 여부 확인 및 생성 함수 추가
    def check_gmail_credentials():
        """Gmail API 인증 정보를 확인합니다"""
        GMAIL_CREDENTIALS_FILE = os.getenv("GMAIL_CREDENTIALS_FILE")
        if not GMAIL_CREDENTIALS_FILE:
            logging.error("GMAIL_CREDENTIALS_FILE 환경 변수가 설정되지 않았습니다.")
            return False

        # 인증 파일이 존재하는지 확인
        if os.path.exists(GMAIL_CREDENTIALS_FILE):
            try:
                # 인증 파일 형식 확인
                with open(GMAIL_CREDENTIALS_FILE, "r") as f:
                    creds_data = json.load(f)

                # OAuth 클라이언트 JSON 파일인 경우 자동 인증 진행 안내
                if "installed" in creds_data or "web" in creds_data:
                    print(
                        f"""
======================================================================
OAuth 클라이언트 인증 파일이 감지되었습니다.
프로그램을 계속 실행하면 브라우저 인증 URL이 표시되며,
안내에 따라 인증 절차를 완료하세요.
======================================================================
                        """
                    )
                    return True

                # 토큰 파일 형식이 아닌 경우
                elif not ("token" in creds_data or "refresh_token" in creds_data):
                    logging.error("인식할 수 없는 인증 파일 형식입니다.")
                    return False

                return True
            except Exception as e:
                logging.error(f"인증 파일 확인 중 오류 발생: {str(e)}")
                return False

        # 인증 파일이 없는 경우
        logging.error(f"Gmail 인증 파일({GMAIL_CREDENTIALS_FILE})이 존재하지 않습니다.")

        # credentials.json 파일 찾기 시도
        credentials_file = find_credentials_file()
        if credentials_file:
            logging.info(f"대체 인증 파일({credentials_file})을 찾았습니다.")
            print(
                f"""
======================================================================
OAuth 클라이언트 인증 파일({credentials_file})이 감지되었습니다.
이 파일을 사용해 Gmail API 인증을 진행할 수 있습니다.
======================================================================
                """
            )
            return True
        else:
            # 인증 파일이 없고, credentials.json도 찾지 못한 경우
            show_credentials_help()
            return False

    # 인증 정보 파일(credentials.json) 찾기 함수
    def find_credentials_file():
        """Google Cloud Console에서 다운로드한 OAuth 클라이언트 JSON 파일(credentials.json)을 찾습니다."""
        GMAIL_CREDENTIALS_FILE = os.getenv("GMAIL_CREDENTIALS_FILE", "")

        # 찾을 파일명 목록 (우선순위 순)
        possible_filenames = [
            "credentials.json",
            "gmail_credentials.json",
            "client_secret.json",
            os.path.join(os.path.dirname(GMAIL_CREDENTIALS_FILE), "credentials.json"),
            os.path.splitext(GMAIL_CREDENTIALS_FILE)[0] + "_credentials.json",
        ]

        # 현재 디렉토리와 하위 디렉토리에서 OAuth 클라이언트 파일 찾기
        for root, dirs, files in os.walk("."):
            for filename in files:
                if filename in ["credentials.json", "client_secret.json"]:
                    file_path = os.path.join(root, filename)
                    try:
                        with open(file_path, "r") as f:
                            data = json.load(f)
                            if "installed" in data or "web" in data:
                                logging.info(
                                    f"유효한 OAuth 클라이언트 파일을 찾았습니다: {file_path}"
                                )
                                return file_path
                    except:
                        pass

        # 지정된 경로에서 OAuth 클라이언트 파일 찾기
        for filename in possible_filenames:
            if os.path.exists(filename):
                try:
                    with open(filename, "r") as f:
                        data = json.load(f)
                        if "installed" in data or "web" in data:
                            logging.info(
                                f"유효한 OAuth 클라이언트 파일을 찾았습니다: {filename}"
                            )
                            return filename
                except:
                    pass

        # 인증 파일을 찾지 못함
        logging.error("유효한 OAuth 클라이언트 파일을 찾지 못했습니다.")
        return None

    # OAuth 클라이언트 파일로 인증하는 함수
    def authenticate_with_oauth(credentials_file):
        """OAuth 클라이언트 JSON 파일을 사용하여 Gmail API 인증 수행"""
        from google_auth_oauthlib.flow import InstalledAppFlow

        try:
            # 인증 파일 로드
            with open(credentials_file, "r") as f:
                client_config = json.load(f)

            # 인증 흐름 생성 - 명시적으로 redirect_uri 설정
            flow = InstalledAppFlow.from_client_config(
                client_config,
                scopes=["https://www.googleapis.com/auth/gmail.send"],
                redirect_uri="urn:ietf:wg:oauth:2.0:oob",
            )

            # 콘솔에서 인증 URL 제공
            auth_url, _ = flow.authorization_url(
                prompt="consent", access_type="offline", include_granted_scopes="true"
            )

            print("\n" + "=" * 70)
            print("Gmail 인증 필요: 아래 URL을 복사하여 브라우저에서 열어주세요")
            print("=" * 70)
            print(auth_url)
            print("=" * 70)
            print(
                "브라우저에서 Google 계정으로 인증 후 나타나는 인증 코드를 복사하여 입력해주세요:"
            )

            # 사용자 인증 코드 입력 대기
            auth_code = input("인증 코드: ").strip()

            # 인증 코드로 토큰 획득
            flow.fetch_token(code=auth_code)

            # 인증 정보 반환
            return flow.credentials

        except Exception as e:
            logging.error(f"OAuth 인증 과정 중 오류 발생: {str(e)}")
            traceback.print_exc()
            return None

    # 인증 관련 도움말 출력 함수
    def show_credentials_help():
        """Gmail 인증 설정 관련 도움말 출력"""
        print(
            f"""
======================================================================
Gmail 인증 파일을 찾을 수 없거나 인증이 만료되었습니다.

OAuth 클라이언트 인증 파일을 준비하는 방법:
1. Google Cloud Console(https://console.cloud.google.com/)에서 프로젝트로 이동
2. API 및 서비스 > 라이브러리에서 Gmail API 활성화
3. API 및 서비스 > 사용자 인증 정보에서 OAuth 클라이언트 ID 생성
   - 애플리케이션 유형: '데스크톱 앱'
   - 승인된 리디렉션 URI에 'urn:ietf:wg:oauth:2.0:oob' 추가
4. OAuth 동의 화면 구성
   - 테스트 사용자로 자신의 계정 추가
   - 필요한 범위(gmail.send) 추가
5. OAuth 클라이언트 ID에서 'JSON 다운로드' 클릭
6. 다운로드한 파일을 이 프로그램의 실행 폴더에 'credentials.json'으로 저장
7. .env 파일에 GMAIL_CREDENTIALS_FILE 값을 'token.json'으로 설정
8. 프로그램을 다시 실행하세요
======================================================================
        """
        )

    # Gmail API를 이용해 다수의 수신자에게 메일 발송 (엑셀 파일 첨부)
    def send_email_with_attachment(subject, body, attachment_filename):
        """Gmail API를 이용해 첨부 파일이 있는 이메일을 발송합니다."""
        GMAIL_SENDER_EMAIL = os.getenv("GMAIL_SENDER_EMAIL")
        GMAIL_RECIPIENTS = os.getenv("GMAIL_RECIPIENTS", "")
        recipient_list = [x.strip() for x in GMAIL_RECIPIENTS.split(",") if x.strip()]
        GMAIL_CREDENTIALS_FILE = os.getenv("GMAIL_CREDENTIALS_FILE")

        if not (GMAIL_CREDENTIALS_FILE and GMAIL_SENDER_EMAIL and recipient_list):
            logging.error("Gmail 발송에 필요한 설정이 누락되었습니다.")
            logging.error(
                "자격증명 파일: %s",
                ("있음" if GMAIL_CREDENTIALS_FILE else "없음"),
            )
            logging.error("발신자 이메일: %s", "있음" if GMAIL_SENDER_EMAIL else "없음")
            logging.error("수신자 목록: %s", "있음" if recipient_list else "없음")
            return False

        try:
            creds = None

            # 인증 토큰 파일 존재 여부 확인 및 유효성 검증
            if not os.path.exists(GMAIL_CREDENTIALS_FILE):
                logging.error(
                    f"Gmail 인증 파일({GMAIL_CREDENTIALS_FILE})이 존재하지 않습니다."
                )

                # credentials.json 파일 찾기 시도
                credentials_file = find_credentials_file()
                if credentials_file:
                    logging.info(
                        f"인증 파일({credentials_file})을 찾았습니다. 인증을 시도합니다."
                    )
                    # 찾은 credentials.json으로 인증
                    creds = authenticate_with_oauth(credentials_file)
                    if creds:
                        logging.info("인증 성공! 새 토큰으로 계속 진행합니다.")
                        # 새 토큰 저장
                        with open(GMAIL_CREDENTIALS_FILE, "w") as token_file:
                            token_data = {
                                "client_id": creds.client_id,
                                "client_secret": creds.client_secret,
                                "refresh_token": creds.refresh_token,
                                "token": creds.token,
                                "token_uri": creds.token_uri,
                                "scopes": creds.scopes,
                            }
                            json.dump(token_data, token_file)
                            logging.info(
                                f"생성된 토큰이 {GMAIL_CREDENTIALS_FILE}에 저장되었습니다."
                            )
                    else:
                        logging.error("인증 실패")
                        show_credentials_help()
                        return False
                else:
                    show_credentials_help()
                    return False

            # 저장된 인증 토큰 읽기 시도
            if not creds:  # 아직 인증 정보가 없는 경우
                try:
                    # 토큰 파일 형식 확인
                    with open(GMAIL_CREDENTIALS_FILE, "r") as f:
                        creds_data = json.load(f)

                    # OAuth 클라이언트 JSON 파일인 경우 - 인증 필요
                    if "installed" in creds_data or "web" in creds_data:
                        logging.info(
                            "새로운 OAuth 클라이언트 인증 파일 감지, 인증 시도..."
                        )
                        creds = authenticate_with_oauth(GMAIL_CREDENTIALS_FILE)
                        if creds:
                            logging.info("인증 성공!")
                            # 새 토큰 저장
                            token_data = {
                                "client_id": creds.client_id,
                                "client_secret": creds.client_secret,
                                "refresh_token": creds.refresh_token,
                                "token": creds.token,
                                "token_uri": creds.token_uri,
                                "scopes": creds.scopes,
                            }
                            # 백업 생성
                            backup_file = f"{GMAIL_CREDENTIALS_FILE}.backup.{datetime.now().strftime('%Y%m%d%H%M%S')}"
                            import shutil

                            shutil.copy2(GMAIL_CREDENTIALS_FILE, backup_file)
                            logging.info(
                                f"OAuth 클라이언트 파일이 {backup_file}으로 백업되었습니다."
                            )

                            # 새 토큰 저장
                            with open(GMAIL_CREDENTIALS_FILE, "w") as token_file:
                                json.dump(token_data, token_file)
                                logging.info(
                                    f"생성된 토큰이 {GMAIL_CREDENTIALS_FILE}에 저장되었습니다."
                                )
                        else:
                            logging.error("OAuth 인증 실패")
                            show_credentials_help()
                            return False

                    # 토큰 파일 형식 확인 (표준 형식)
                    elif "token" in creds_data or "refresh_token" in creds_data:
                        creds = Credentials.from_authorized_user_info(
                            creds_data,
                            scopes=["https://www.googleapis.com/auth/gmail.send"],
                        )
                        logging.info("기존 인증 토큰을 성공적으로 로드했습니다.")
                    else:
                        logging.error("인식할 수 없는 인증 파일 형식입니다.")
                        return False

                except Exception as token_err:
                    logging.error(f"자격증명 로드 실패: {str(token_err)}")
                    return False

            # 토큰 만료 시 갱신 시도
            if creds and (creds.expired or not creds.valid):
                logging.error("만료된 토큰 갱신 시도 중...")
                try:
                    creds.refresh(Request())
                    # 갱신된 토큰 저장
                    token_data = {
                        "client_id": creds.client_id,
                        "client_secret": creds.client_secret,
                        "refresh_token": creds.refresh_token,
                        "token": creds.token,
                        "token_uri": creds.token_uri,
                        "scopes": creds.scopes,
                    }

                    # 백업 생성
                    backup_file = f"{GMAIL_CREDENTIALS_FILE}.backup.{datetime.now().strftime('%Y%m%d%H%M%S')}"
                    import shutil

                    if os.path.exists(GMAIL_CREDENTIALS_FILE):
                        shutil.copy2(GMAIL_CREDENTIALS_FILE, backup_file)
                        logging.error(
                            f"기존 토큰 파일이 {backup_file}으로 백업되었습니다."
                        )

                    # 갱신된 토큰 저장
                    with open(GMAIL_CREDENTIALS_FILE, "w") as token_file:
                        json.dump(token_data, token_file)
                        logging.error("갱신된 토큰이 파일에 저장되었습니다.")

                except Exception as refresh_err:
                    logging.error(f"토큰 갱신 실패: {str(refresh_err)}")
                    traceback.print_exc()

                    # invalid_grant 또는 deleted_client 에러 처리
                    if (
                        "invalid_grant" in str(refresh_err).lower()
                        or "deleted_client" in str(refresh_err).lower()
                    ):
                        logging.error(
                            "클라이언트가 삭제되었거나 인증이 만료되었습니다. 자동 재인증을 시도합니다."
                        )

                        # credentials.json 파일 찾기 시도
                        credentials_file = find_credentials_file()
                        if credentials_file:
                            logging.error(
                                f"인증 파일({credentials_file})을 찾았습니다. 재인증을 시도합니다."
                            )

                            # 찾은 credentials.json으로 재인증
                            try:
                                new_creds = authenticate_with_oauth(credentials_file)
                                if new_creds:
                                    logging.error(
                                        "재인증 성공! 메일 발송을 재시도합니다."
                                    )

                                    # 갱신된 토큰 저장
                                    token_data = {
                                        "client_id": new_creds.client_id,
                                        "client_secret": new_creds.client_secret,
                                        "refresh_token": new_creds.refresh_token,
                                        "token": new_creds.token,
                                        "token_uri": new_creds.token_uri,
                                        "scopes": new_creds.scopes,
                                    }
                                    with open(
                                        GMAIL_CREDENTIALS_FILE, "w"
                                    ) as token_file:
                                        json.dump(token_data, token_file)
                                        logging.error(
                                            f"갱신된 토큰이 {GMAIL_CREDENTIALS_FILE}에 저장되었습니다."
                                        )

                                    # 메일 재전송 시도 - 새로운 인증 정보로 다시 시도
                                    return send_email_with_attachment(
                                        subject, body, attachment_filename
                                    )
                                else:
                                    logging.error("재인증 실패")
                                    show_credentials_help()
                            except Exception as auth_err:
                                logging.error(f"재인증 중 예외 발생: {str(auth_err)}")
                                traceback.print_exc()
                                show_credentials_help()
                        else:
                            logging.error("인증 파일을 찾지 못했습니다.")
                            show_credentials_help()
                except Exception as backup_err:
                    logging.error(f"토큰 파일 백업 중 오류 발생: {str(backup_err)}")
                    traceback.print_exc()
                    show_credentials_help()

            # 인증 정보 유효성 검증
            if not creds or not creds.valid:
                logging.error("유효한 자격증명을 얻지 못했습니다.")
                return False

            # Gmail API 서비스 구축
            service = build("gmail", "v1", credentials=creds)

            # 이메일 메시지 생성
            message = MIMEMultipart()
            message["to"] = ", ".join(recipient_list)
            message["from"] = GMAIL_SENDER_EMAIL
            message["subject"] = subject
            message.attach(MIMEText(body, "plain"))

            # 첨부 파일 추가
            try:
                with open(attachment_filename, "rb") as f:
                    attachment_data = f.read()
                attachment_part = MIMEApplication(
                    attachment_data, Name=os.path.basename(attachment_filename)
                )
                attachment_part["Content-Disposition"] = (
                    f'attachment; filename="{os.path.basename(attachment_filename)}"'
                )
                message.attach(attachment_part)
            except Exception as attach_err:
                logging.error(f"첨부파일 처리 중 오류 발생: {str(attach_err)}")
                traceback.print_exc()
                return False

            # 메시지 인코딩 및 전송
            raw_message = base64.urlsafe_b64encode(message.as_bytes()).decode()
            message_body = {"raw": raw_message}

            sent_message = (
                service.users()
                .messages()
                .send(userId="me", body=message_body)
                .execute()
            )
            logging.info(f"이메일 발송 성공: {sent_message['id']}")
            return True

        except Exception as e:
            logging.error(f"Gmail 발송 중 오류 발생: {str(e)}")
            traceback.print_exc()

            # invalid_grant 에러 특별 처리
            if "invalid_grant" in str(e).lower() or "deleted_client" in str(e).lower():
                logging.error("인증 토큰이 만료되었습니다. 자동 갱신을 시도합니다.")

                # 토큰 백업 생성
                backup_file = f"{GMAIL_CREDENTIALS_FILE}.backup.{datetime.now().strftime('%Y%m%d%H%M%S')}"
                try:
                    import shutil

                    # 기존 파일 백업
                    if os.path.exists(GMAIL_CREDENTIALS_FILE):
                        shutil.copy2(GMAIL_CREDENTIALS_FILE, backup_file)
                        logging.error(
                            f"기존 토큰 파일이 {backup_file}으로 백업되었습니다."
                        )

                    # credentials.json 파일 찾기 시도
                    credentials_file = find_credentials_file()
                    if credentials_file:
                        logging.error(
                            f"인증 파일({credentials_file})을 찾았습니다. 재인증을 시도합니다."
                        )

                        # 찾은 credentials.json으로 재인증
                        try:
                            new_creds = authenticate_with_oauth(credentials_file)
                            if new_creds:
                                logging.error("재인증 성공! 메일 발송을 재시도합니다.")

                                # 갱신된 토큰 저장
                                token_data = {
                                    "client_id": new_creds.client_id,
                                    "client_secret": new_creds.client_secret,
                                    "refresh_token": new_creds.refresh_token,
                                    "token": new_creds.token,
                                    "token_uri": new_creds.token_uri,
                                    "scopes": new_creds.scopes,
                                }
                                with open(GMAIL_CREDENTIALS_FILE, "w") as token_file:
                                    json.dump(token_data, token_file)
                                    logging.error(
                                        f"갱신된 토큰이 {GMAIL_CREDENTIALS_FILE}에 저장되었습니다."
                                    )

                                # 메일 재전송 시도 - 새로운 인증 정보로 다시 시도
                                return send_email_with_attachment(
                                    subject, body, attachment_filename
                                )
                            else:
                                logging.error("재인증 실패")
                                show_credentials_help()
                        except Exception as auth_err:
                            logging.error(f"재인증 중 예외 발생: {str(auth_err)}")
                            traceback.print_exc()
                            show_credentials_help()
                    else:
                        logging.error("인증 파일을 찾지 못했습니다.")
                        show_credentials_help()
                except Exception as backup_err:
                    logging.error(f"토큰 파일 백업 중 오류 발생: {str(backup_err)}")
                    traceback.print_exc()
                    show_credentials_help()

            return False

    # 15. 메일 발송 정보 구성
    today_str = datetime.now().strftime("%Y-%m-%d")
    email_subject = f"YNA 뉴스 결과 - {today_str}"
    email_body = (
        f"오늘 날짜: {today_str}\n"
        f"검색 키워드: {', '.join(keyword_list)}\n"
        f"총 뉴스 기사 수: {len(filtered_articles)}개\n\n"
        f"OpenAI API 사용량:\n"
        f"- 입력 토큰: {usage_data['input_tokens']:,}개 (${usage_data['input_cost_usd']:.6f}, ₩{usage_data['input_cost_krw']:.2f})\n"
        f"- 출력 토큰: {usage_data['output_tokens']:,}개 (${usage_data['output_cost_usd']:.6f}, ₩{usage_data['output_cost_krw']:.2f})\n"
        f"- 전체 토큰: {usage_data['total_tokens']:,}개 (${usage_data['total_cost_usd']:.6f}, ₩{usage_data['total_cost_krw']:.2f})\n\n"
        "첨부된 엑셀 파일을 확인해 주세요."
    )

    # 16. 메일 발송 (환경변수 SEND_EMAIL이 true인 경우)
    if os.getenv("SEND_EMAIL", "").lower() == "true":
        logging.info("이메일 발송 시도 중...")

        try:
            # Gmail 발송에 필요한 설정 확인
            GMAIL_SENDER_EMAIL = os.getenv("GMAIL_SENDER_EMAIL")
            GMAIL_RECIPIENTS = os.getenv("GMAIL_RECIPIENTS", "")
            recipient_list = [
                x.strip() for x in GMAIL_RECIPIENTS.split(",") if x.strip()
            ]

            if not (GMAIL_SENDER_EMAIL and recipient_list):
                logging.error("Gmail 발송에 필요한 환경변수가 설정되지 않았습니다.")
                if not GMAIL_SENDER_EMAIL:
                    logging.error("GMAIL_SENDER_EMAIL 환경변수가 설정되지 않았습니다.")
                if not recipient_list:
                    logging.error(
                        "GMAIL_RECIPIENTS 환경변수가 설정되지 않았거나 유효한 이메일이 없습니다."
                    )
                print("Gmail 설정을 확인하세요. 자세한 내용은 로그 파일을 참조하세요.")
            # Gmail 인증 정보 확인
            elif not check_gmail_credentials():
                logging.error(
                    "Gmail 인증 정보가 유효하지 않아 이메일을 발송할 수 없습니다."
                )
                print(
                    "Gmail 인증 정보를 확인하세요. 자세한 내용은 로그 파일을 참조하세요."
                )
            else:
                # 이메일 발송
                email_sent = send_email_with_attachment(
                    email_subject, email_body, output_filename
                )

                if email_sent:
                    logging.info(f"이메일이 성공적으로 발송되었습니다: {email_subject}")
                    print(f"이메일 발송 완료 - 수신자: {GMAIL_RECIPIENTS}")
                else:
                    logging.error("이메일 발송에 실패했습니다.")
                    print(
                        "이메일 발송에 실패했습니다. 자세한 내용은 로그 파일을 참조하세요."
                    )
        except Exception as e:
            logging.error(
                f"이메일 발송 과정에서 예상치 못한 오류가 발생했습니다: {str(e)}"
            )
            traceback.print_exc()
            print(
                "이메일 발송 중 오류가 발생했습니다. 자세한 내용은 로그 파일을 참조하세요."
            )
    else:
        logging.info("이메일 발송 기능이 비활성화되어 있습니다. (SEND_EMAIL=false)")

    # 최종 파일 정보 출력
    logging.info("==== 결과 파일 정보 ====")
    logging.info(f"엑셀 파일: {os.path.abspath(output_filename)}")
    logging.info(f"로그 파일: {os.path.abspath(log_path)}")
    logging.info("======================")
    logging.info("프로그램 종료")


if __name__ == "__main__":
    main()
